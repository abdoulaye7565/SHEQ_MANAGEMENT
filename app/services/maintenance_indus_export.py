"""maintenance_indus_export.py — Export XLSX (openpyxl) du module Maintenance Industrielle.

Génère un classeur identique au template maintenance_equipements_industriels.py :
7 feuilles, palette Navy/Gold, mise en forme conditionnelle.
"""
from __future__ import annotations

import logging
from datetime import date, timedelta
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from app.services.maintenance_indus_service import (
    get_indicateurs_mtbf,
    get_maintenance_dashboard,
    list_equipements,
    list_interventions,
    list_pieces,
    list_plan_pm,
    list_prestataires,
)

_log = logging.getLogger(__name__)

# ── Palette (identique au template) ──────────────────────────────────────────
NAVY        = "0D1B2A"
STEEL       = "1C3557"
GOLD        = "C8A400"
GREEN       = "00A86B"
RED         = "C0392B"
BLUE        = "2E86C1"
ORANGE      = "E67E22"
TEAL        = "148F77"
PURPLE      = "6C3483"
WHITE       = "FFFFFF"
LIGHT_GRAY  = "F2F3F4"
MED_GRAY    = "D5D8DC"
DARK_GRAY   = "2C3E50"
LIGHT_BLUE  = "D6EAF8"
LIGHT_GREEN = "D5F5E3"
LIGHT_RED   = "FADBD8"
LIGHT_ORANGE= "FDEBD0"

TODAY = date.today()
YEAR  = TODAY.year


# ── Helpers style ─────────────────────────────────────────────────────────────
def _fill(hex_c: str) -> PatternFill:
    return PatternFill("solid", fgColor=hex_c)

def _font(bold=False, size=10, color=DARK_GRAY, italic=False) -> Font:
    return Font(bold=bold, size=size, color=color, italic=italic, name="Calibri")

def _align(h="left", v="center", wrap=False) -> Alignment:
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

def _border() -> Border:
    t = Side(style="thin", color=MED_GRAY)
    return Border(left=t, right=t, top=t, bottom=t)

def _col_w(ws, widths: dict) -> None:
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

def _title_row(ws, row: int, text: str, ncols: int,
               bg: str = NAVY, fg: str = GOLD, size: int = 14) -> None:
    last_col = get_column_letter(ncols)
    ws.merge_cells(f"A{row}:{last_col}{row}")
    c = ws[f"A{row}"]
    c.value = text
    c.fill = _fill(bg)
    c.font = Font(bold=True, size=size, color=fg, name="Calibri")
    c.alignment = _align(h="center")
    ws.row_dimensions[row].height = 36 if size >= 14 else 22

def _sub_row(ws, row: int, text: str, ncols: int,
             bg: str = STEEL, fg: str = WHITE) -> None:
    last_col = get_column_letter(ncols)
    ws.merge_cells(f"A{row}:{last_col}{row}")
    c = ws[f"A{row}"]
    c.value = text
    c.fill = _fill(bg)
    c.font = Font(bold=False, size=9, color=fg, italic=True, name="Calibri")
    c.alignment = _align(h="center")
    ws.row_dimensions[row].height = 16

def _header(ws, row: int, labels: list[str],
            bg: str = STEEL, fg: str = GOLD) -> None:
    for i, lbl in enumerate(labels, 1):
        c = ws.cell(row=row, column=i, value=lbl)
        c.fill = _fill(bg)
        c.font = Font(bold=True, size=9, color=fg, name="Calibri")
        c.alignment = _align(h="center", v="center", wrap=True)
        c.border = _border()
    ws.row_dimensions[row].height = 28

def _cell(ws, row: int, col: int, value, bg: str = WHITE,
          fg: str = DARK_GRAY, bold: bool = False,
          h: str = "left", wrap: bool = False, fmt: str = "") -> None:
    c = ws.cell(row=row, column=col, value=value)
    c.fill = _fill(bg)
    c.font = Font(bold=bold, size=9, color=fg, name="Calibri")
    c.alignment = _align(h=h, v="center", wrap=wrap)
    c.border = _border()
    if fmt:
        c.number_format = fmt
    ws.row_dimensions[row].height = 17

def _kpi_block(ws, row: int, col: int, title: str, value: str,
               subtitle: str, bg_title: str, bg_val: str) -> None:
    """3 lignes × 3 colonnes."""
    c1 = get_column_letter(col)
    c3 = get_column_letter(col + 2)
    ws.merge_cells(f"{c1}{row}:{c3}{row}")
    t = ws[f"{c1}{row}"]
    t.value = title; t.fill = _fill(bg_title)
    t.font = Font(bold=True, size=9, color=WHITE, name="Calibri")
    t.alignment = _align(h="center"); ws.row_dimensions[row].height = 18

    ws.merge_cells(f"{c1}{row+1}:{c3}{row+1}")
    v = ws[f"{c1}{row+1}"]
    v.value = value; v.fill = _fill(bg_val)
    v.font = Font(bold=True, size=18, color=GOLD, name="Calibri")
    v.alignment = _align(h="center"); ws.row_dimensions[row+1].height = 30

    ws.merge_cells(f"{c1}{row+2}:{c3}{row+2}")
    s = ws[f"{c1}{row+2}"]
    s.value = subtitle; s.fill = _fill(STEEL)
    s.font = Font(size=8, color=WHITE, italic=True, name="Calibri")
    s.alignment = _align(h="center"); ws.row_dimensions[row+2].height = 14


# ── Couleurs par catégorie ─────────────────────────────────────────────────────
CRIT_BG  = {"A": LIGHT_RED,    "B": LIGHT_ORANGE, "C": "D1F2EB"}
CRIT_FG  = {"A": RED,           "B": ORANGE,        "C": TEAL}
STAT_EQ  = {"En service": (LIGHT_GREEN, GREEN), "Maintenance": (LIGHT_ORANGE, ORANGE),
             "Arret": (LIGHT_RED, RED), "Reforme": (LIGHT_GRAY, DARK_GRAY)}
FREQ_BG  = {"Journalier": "E8F8F5", "Hebdomadaire": "EBF5FB", "Mensuel": "F5EEF8",
             "Trimestriel": LIGHT_ORANGE, "Semestriel": LIGHT_RED, "Annuel": LIGHT_GRAY}
FREQ_FG  = {"Journalier": TEAL, "Hebdomadaire": BLUE, "Mensuel": PURPLE,
             "Trimestriel": ORANGE, "Semestriel": RED, "Annuel": DARK_GRAY}
TYPE_BG  = {"Préventive systématique": LIGHT_GREEN, "Préventive conditionnelle": "D1F2EB",
             "Corrective urgente": LIGHT_RED,       "Corrective planifiée": LIGHT_ORANGE,
             "Améliorative": "E8DAEF",              "Prédictive": LIGHT_BLUE}
TYPE_FG  = {"Préventive systématique": GREEN, "Préventive conditionnelle": TEAL,
             "Corrective urgente": RED,       "Corrective planifiée": ORANGE,
             "Améliorative": PURPLE,          "Prédictive": BLUE}
STAT_OT  = {"Cloture": (LIGHT_GREEN, GREEN), "En cours": (LIGHT_BLUE, BLUE),
             "Ouvert": (LIGHT_ORANGE, ORANGE)}


def _unique_path(base: str) -> Path:
    try:
        from app.services.attendance_export_service import _unique_export_path
        return _unique_export_path(base)
    except ImportError:
        p = Path(base)
        if not p.exists():
            return p
        for i in range(1, 999):
            q = p.with_name(f"{p.stem}_{i}{p.suffix}")
            if not q.exists():
                return q
        return p


# ═══════════════════════════════════════════════════════════════════════════════
def export_maintenance_xlsx() -> Path:
    path = _unique_path(f"maintenance_industrielle_{YEAR}.xlsx")

    # Charger données
    try: dash = get_maintenance_dashboard()
    except Exception: dash = {}
    try: equipements = list_equipements()
    except Exception: equipements = []
    try: plan_pm = list_plan_pm()
    except Exception: plan_pm = []
    try: interventions = list_interventions()
    except Exception: interventions = []
    try: pieces = list_pieces()
    except Exception: pieces = []
    try: prestataires = list_prestataires()
    except Exception: prestataires = []
    try: indicateurs = get_indicateurs_mtbf()
    except Exception: indicateurs = []

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    _sheet_dashboard(wb, dash, equipements, pieces)
    _sheet_equipements(wb, equipements)
    _sheet_plan_pm(wb, plan_pm)
    _sheet_interventions(wb, interventions)
    _sheet_pieces(wb, pieces)
    _sheet_prestataires(wb, prestataires)
    _sheet_indicateurs(wb, indicateurs, dash)

    wb.save(str(path))
    _log.info("[maint_export] -> %s", path)
    return path


# ── FEUILLE 1 : DASHBOARD ─────────────────────────────────────────────────────
def _sheet_dashboard(wb, dash: dict, equipements: list, pieces: list) -> None:
    ws = wb.create_sheet("DASHBOARD")
    ws.sheet_view.showGridLines = False
    ws.sheet_properties.tabColor = GOLD

    nb_eq      = dash.get("nb_equipements", len(equipements))
    nb_actif   = dash.get("nb_eq_actif", 0)
    nb_ot_ytd  = dash.get("nb_ot_ytd", 0)
    nb_pm      = dash.get("nb_pm_ytd", 0)
    nb_cm      = dash.get("nb_cm_ytd", 0)
    cout_ytd   = dash.get("cout_ytd", 0)
    h_arret    = dash.get("h_arret_ytd", 0)
    ratio_pm   = dash.get("ratio_pm", 0)
    nb_alertes = dash.get("nb_alerte_pieces", 0)
    val_parc   = sum(float(e.get("valeur_remplacement") or 0) for e in equipements)

    _title_row(ws, 1,
               f"MAINTENANCE INDUSTRIELLE — DIAMOND DRILLING  |  {YEAR}", 18)
    _sub_row(ws, 2,
             f"Export du {TODAY.strftime('%d/%m/%Y')}   |   "
             f"Équipements : {nb_eq}   |   OT YTD : {nb_ot_ytd}   |   "
             f"Valeur parc : {val_parc:,.0f} FCFA", 18)
    ws.row_dimensions[3].height = 8

    # KPIs (2 lignes de 3 blocs)
    _kpi_block(ws, 4, 1,  "ÉQUIPEMENTS ACTIFS",   f"{nb_actif}/{nb_eq}",    "En service / Total",           NAVY,  NAVY)
    _kpi_block(ws, 4, 4,  "ORDRES DE TRAVAIL YTD",str(nb_ot_ytd),           f"PM:{nb_pm}  CM:{nb_cm}",       STEEL, STEEL)
    _kpi_block(ws, 4, 7,  "COÛT MAINTENANCE YTD", f"{cout_ytd/1e6:.1f} M",  "Millions FCFA",                RED,   DARK_GRAY)
    _kpi_block(ws, 4, 10, "ARRÊTS PANNES YTD",    f"{h_arret:.0f} h",       "Heures d'immobilisation",      RED,   DARK_GRAY)
    _kpi_block(ws, 4, 13, "RATIO PM / CM",         f"{ratio_pm:.0f}%",       "Objectif ≥ 70% préventif",     TEAL,  TEAL)
    _kpi_block(ws, 4, 16, "ALERTES PIÈCES",        str(nb_alertes),          "Stock sous seuil minimum",
               RED if nb_alertes > 0 else GREEN, RED if nb_alertes > 0 else GREEN)
    ws.row_dimensions[7].height = 8

    # Tableau équipements critiques (A)
    eq_a = [e for e in equipements if str(e.get("criticite","")) == "A"]
    _title_row(ws, 8, "ÉQUIPEMENTS CRITIQUES — CRITICITÉ A", 18, bg=STEEL, size=11)
    ws.row_dimensions[8].height = 22
    hdr = ["Code", "Désignation", "Famille", "Zone", "Marque", "Date MES",
           "Statut", "Valeur remplacement (FCFA)"]
    _header(ws, 9, hdr, bg=NAVY, fg=GOLD)
    for i, e in enumerate(eq_a):
        r = 10 + i
        bg = LIGHT_BLUE if i % 2 == 0 else WHITE
        stat = str(e.get("statut","En service"))
        sbg, sfg = STAT_EQ.get(stat, (LIGHT_GRAY, DARK_GRAY))
        _cell(ws, r, 1, e.get("code_equipement",""), bg=bg, bold=True, fg=BLUE, h="center")
        _cell(ws, r, 2, e.get("designation",""), bg=bg)
        _cell(ws, r, 3, e.get("famille",""), bg=bg)
        _cell(ws, r, 4, e.get("site_zone",""), bg=bg)
        _cell(ws, r, 5, e.get("marque",""), bg=bg)
        _cell(ws, r, 6, e.get("date_mise_en_service",""), bg=bg, h="center")
        _cell(ws, r, 7, stat, bg=sbg, fg=sfg, bold=True, h="center")
        _cell(ws, r, 8, float(e.get("valeur_remplacement") or 0),
              bg=bg, fg=DARK_GRAY, h="right", fmt="#,##0")

    gap = 10 + len(eq_a) + 1
    ws.row_dimensions[gap].height = 8

    # Alertes pièces
    alertes = [p for p in pieces
               if float(p.get("stock_actuel") or 0) <= float(p.get("stock_min") or 0)]
    _title_row(ws, gap + 1, "ALERTES PIÈCES DE RECHANGE — STOCK SOUS SEUIL", 18,
               bg=RED, size=11)
    ws.row_dimensions[gap + 1].height = 22
    _header(ws, gap + 2,
            ["Criticité","Code","Désignation","Stock actuel","Stock min",
             "Emplacement","Délai appro (j)","Fournisseur"], bg=RED, fg=WHITE)
    for i, p in enumerate(alertes):
        r = gap + 3 + i
        crit = str(p.get("criticite","B"))
        cbg, cfg = CRIT_BG.get(crit, (LIGHT_GRAY, DARK_GRAY)), CRIT_FG.get(crit, DARK_GRAY)
        bg = LIGHT_RED if float(p.get("stock_actuel") or 0) == 0 else LIGHT_ORANGE
        _cell(ws, r, 1, crit,                            bg=cbg, fg=cfg, bold=True, h="center")
        _cell(ws, r, 2, p.get("code_piece",""),           bg=bg, fg=DARK_GRAY, bold=True)
        _cell(ws, r, 3, p.get("designation",""),          bg=bg)
        _cell(ws, r, 4, float(p.get("stock_actuel") or 0),bg=LIGHT_RED, fg=RED, bold=True, h="right", fmt="#,##0.0")
        _cell(ws, r, 5, float(p.get("stock_min") or 0),  bg=bg, h="right", fmt="#,##0.0")
        _cell(ws, r, 6, p.get("emplacement_magasin",""),  bg=bg)
        _cell(ws, r, 7, int(p.get("delai_appro") or 0),   bg=bg, h="center")
        _cell(ws, r, 8, p.get("fournisseur",""),           bg=bg)

    if not alertes:
        r = gap + 3
        ws.merge_cells(f"A{r}:H{r}")
        c = ws[f"A{r}"]
        c.value = "✓  Aucune alerte — tous les stocks sont suffisants"
        c.fill = _fill(LIGHT_GREEN); c.font = Font(size=10, color=GREEN, name="Calibri")
        c.alignment = _align(h="center")

    _col_w(ws, {
        "A":12,"B":12,"C":28,"D":18,"E":18,"F":12,"G":14,
        "H":18,"I":16,"J":14,"K":12,"L":12,"M":16,"N":14,"O":14,"P":14,"Q":14,"R":14,
    })
    ws.freeze_panes = "A3"


# ── FEUILLE 2 : ÉQUIPEMENTS ───────────────────────────────────────────────────
def _sheet_equipements(wb, equipements: list) -> None:
    ws = wb.create_sheet("ÉQUIPEMENTS")
    ws.sheet_view.showGridLines = False
    ws.sheet_properties.tabColor = BLUE

    val_parc = sum(float(e.get("valeur_remplacement") or 0) for e in equipements)
    na, nb_c, nc = (sum(1 for e in equipements if e.get("criticite")==c) for c in "ABC")

    _title_row(ws, 1, f"REGISTRE DES ÉQUIPEMENTS INDUSTRIELS — {YEAR}", 16)
    _sub_row(ws, 2,
             f"{len(equipements)} équipements   |   "
             f"Criticité A:{na}  B:{nb_c}  C:{nc}   |   "
             f"Valeur parc : {val_parc:,.0f} FCFA", 16)
    ws.row_dimensions[3].height = 6

    hdr = ["Code","Désignation","Famille","Sous-famille","Criticité","Site / Zone",
           "Emplacement","Marque","Modèle","N° Série","Date MES",
           "Capacité","Fournisseur","Valeur remplacement (FCFA)","Statut","Observations"]
    _header(ws, 4, hdr, bg=NAVY, fg=GOLD)

    for i, e in enumerate(equipements):
        r = 5 + i
        bg = LIGHT_BLUE if i % 2 == 0 else WHITE
        crit = str(e.get("criticite","B"))
        cbg, cfg = CRIT_BG.get(crit, (LIGHT_GRAY, DARK_GRAY)), CRIT_FG.get(crit, DARK_GRAY)
        stat = str(e.get("statut","En service"))
        sbg, sfg = STAT_EQ.get(stat, (LIGHT_GRAY, DARK_GRAY))
        _cell(ws, r,  1, e.get("code_equipement",""), bg=bg, bold=True, fg=BLUE)
        _cell(ws, r,  2, e.get("designation",""),     bg=bg)
        _cell(ws, r,  3, e.get("famille",""),         bg=bg)
        _cell(ws, r,  4, e.get("sous_famille",""),    bg=bg)
        _cell(ws, r,  5, crit,  bg=cbg, fg=cfg, bold=True, h="center")
        _cell(ws, r,  6, e.get("site_zone",""),        bg=bg)
        _cell(ws, r,  7, e.get("emplacement",""),      bg=bg)
        _cell(ws, r,  8, e.get("marque",""),           bg=bg)
        _cell(ws, r,  9, e.get("modele",""),           bg=bg)
        _cell(ws, r, 10, e.get("numero_serie",""),     bg=bg)
        _cell(ws, r, 11, e.get("date_mise_en_service",""), bg=bg, h="center")
        _cell(ws, r, 12, e.get("capacite_puissance",""),   bg=bg, h="center")
        _cell(ws, r, 13, e.get("fournisseur",""),      bg=bg)
        _cell(ws, r, 14, float(e.get("valeur_remplacement") or 0),
              bg=bg, fg=DARK_GRAY, h="right", bold=True, fmt="#,##0")
        _cell(ws, r, 15, stat, bg=sbg, fg=sfg, bold=True, h="center")
        _cell(ws, r, 16, e.get("observations",""), bg=bg, wrap=True)

    # Ligne totaux
    last = 4 + len(equipements) + 1
    ws.merge_cells(f"A{last}:M{last}")
    c = ws[f"A{last}"]
    c.value = "VALEUR TOTALE PARC"
    c.fill = _fill(GOLD); c.font = Font(bold=True, size=10, color=NAVY, name="Calibri")
    c.alignment = _align(h="right")
    _cell(ws, last, 14, val_parc, bg=GOLD, fg=NAVY, bold=True, h="right", fmt="#,##0")

    _col_w(ws, {"A":14,"B":30,"C":20,"D":20,"E":9,"F":18,"G":16,
                "H":14,"I":14,"J":16,"K":12,"L":18,"M":20,"N":24,"O":14,"P":25})
    ws.freeze_panes = "B5"
    ws.auto_filter.ref = f"A4:P{last-1}"


# ── FEUILLE 3 : PLAN PM ───────────────────────────────────────────────────────
def _sheet_plan_pm(wb, plan_pm: list) -> None:
    ws = wb.create_sheet("PLAN PM")
    ws.sheet_view.showGridLines = False
    ws.sheet_properties.tabColor = GREEN

    today_iso = TODAY.isoformat()
    nb_echus  = sum(1 for p in plan_pm
                    if p.get("prochaine_echeance") and
                    str(p["prochaine_echeance"]) < today_iso)

    _title_row(ws, 1, f"PLAN DE MAINTENANCE PRÉVENTIVE — {YEAR}", 14)
    _sub_row(ws, 2,
             f"{len(plan_pm)} tâches PM   |   {nb_echus} tâches échues   |   "
             "J:Journalier  S:Hebdomadaire  M:Mensuel  T:Trimestriel  A:Annuel", 14)
    ws.row_dimensions[3].height = 6

    hdr = ["Code Plan","Code Équip.","Désignation","Tâche","Fréquence",
           "Dernière réalisation","Prochaine échéance","Durée (h)",
           "Ressources","Pièces nécessaires","Instructions","Responsable","Statut"]
    _header(ws, 4, hdr, bg=NAVY, fg=GREEN)

    for i, p in enumerate(plan_pm):
        r = 5 + i
        proch   = str(p.get("prochaine_echeance") or "")
        echu    = proch and proch < today_iso
        bg      = LIGHT_RED if echu else (LIGHT_BLUE if i % 2 == 0 else WHITE)
        freq    = str(p.get("frequence",""))
        fbg     = FREQ_BG.get(freq, LIGHT_GRAY)
        ffg     = FREQ_FG.get(freq, DARK_GRAY)
        statut  = "ÉCHU" if echu else "OK"
        sbg, sfg = (LIGHT_RED, RED) if echu else (LIGHT_GREEN, GREEN)

        # Calcul code plan si absent
        code_plan = str(p.get("id_plan") or "")
        if code_plan and not code_plan.startswith("PM"):
            code_plan = f"PM-{code_plan:0>3}"

        _cell(ws, r,  1, code_plan,                        bg=bg, fg=DARK_GRAY, bold=True)
        _cell(ws, r,  2, p.get("code_equipement",""),      bg=bg, fg=BLUE)
        _cell(ws, r,  3, p.get("designation_eq",""),       bg=bg)
        _cell(ws, r,  4, p.get("tache",""),                bg=bg, wrap=True)
        _cell(ws, r,  5, freq,                             bg=fbg, fg=ffg, bold=True, h="center")
        _cell(ws, r,  6, str(p.get("derniere_realisation","") or ""),  bg=bg, h="center")
        _cell(ws, r,  7, proch,                            bg=LIGHT_RED if echu else bg,
              fg=RED if echu else DARK_GRAY, bold=echu, h="center")
        _cell(ws, r,  8, float(p.get("duree_h") or 0),    bg=bg, h="center", fmt="0.0")
        _cell(ws, r,  9, p.get("ressources",""),           bg=bg, wrap=True)
        _cell(ws, r, 10, p.get("pieces_necessaires",""),   bg=bg, wrap=True)
        _cell(ws, r, 11, p.get("instructions",""),         bg=bg, wrap=True)
        _cell(ws, r, 12, p.get("responsable",""),          bg=bg)
        _cell(ws, r, 13, statut, bg=sbg, fg=sfg, bold=True, h="center")

    _col_w(ws, {"A":10,"B":12,"C":24,"D":35,"E":14,"F":18,"G":18,
                "H":9,"I":22,"J":28,"K":40,"L":20,"M":10})
    ws.freeze_panes = "D5"
    ws.auto_filter.ref = f"A4:M{4+len(plan_pm)}"


# ── FEUILLE 4 : INTERVENTIONS ─────────────────────────────────────────────────
def _sheet_interventions(wb, interventions: list) -> None:
    ws = wb.create_sheet("INTERVENTIONS")
    ws.sheet_view.showGridLines = False
    ws.sheet_properties.tabColor = ORANGE

    t_arret = sum(float(o.get("temps_arret") or 0) for o in interventions)
    t_mo    = sum(float(o.get("cout_mo") or 0) for o in interventions)
    t_pi    = sum(float(o.get("cout_pieces") or 0) for o in interventions)
    t_tot   = sum(float(o.get("cout_total") or 0) for o in interventions)

    _title_row(ws, 1, f"REGISTRE DES INTERVENTIONS (ORDRES DE TRAVAIL) — {YEAR}", 17)
    _sub_row(ws, 2,
             f"{len(interventions)} OT   |   "
             f"Coût MO : {t_mo:,.0f} F   |   Pièces : {t_pi:,.0f} F   |   "
             f"Total : {t_tot:,.0f} F   |   Arrêts : {t_arret:.0f} h", 17)
    ws.row_dimensions[3].height = 6

    hdr = ["N° OT","Date ouverture","Date clôture","Code Équip.","Désignation",
           "Type","Nature panne","Description travaux","Technicien",
           "Arrêt (h)","Durée interv. (h)","Pièces utilisées",
           "Coût MO (FCFA)","Coût pièces (FCFA)","Coût total (FCFA)","Statut","Observations"]
    _header(ws, 4, hdr, bg=NAVY, fg=ORANGE)

    for i, o in enumerate(interventions):
        r   = 5 + i
        t   = str(o.get("type_maintenance",""))
        tbg = TYPE_BG.get(t, LIGHT_GRAY)
        tfg = TYPE_FG.get(t, DARK_GRAY)
        st  = str(o.get("statut",""))
        sbg, sfg = STAT_OT.get(st, (LIGHT_GRAY, DARK_GRAY))
        arret = float(o.get("temps_arret") or 0)
        cout  = float(o.get("cout_total") or 0)
        # Ligne rouge si CM urgente en cours
        bg = LIGHT_RED if ("urgente" in t and "cours" in st) else (
            LIGHT_BLUE if i % 2 == 0 else WHITE)

        _cell(ws, r,  1, o.get("numero_ot",""),              bg=bg, fg=tfg, bold=True)
        _cell(ws, r,  2, str(o.get("date_ouverture",""))[:10], bg=bg, h="center")
        _cell(ws, r,  3, str(o.get("date_cloture","") or "")[:10] or "—", bg=bg, h="center")
        _cell(ws, r,  4, o.get("code_equipement",""),         bg=bg, fg=BLUE)
        _cell(ws, r,  5, o.get("designation_eq",""),          bg=bg)
        _cell(ws, r,  6, t,  bg=tbg, fg=tfg, bold=True, h="center", wrap=True)
        _cell(ws, r,  7, o.get("nature_panne",""),            bg=bg, wrap=True)
        _cell(ws, r,  8, o.get("description_travaux",""),     bg=bg, wrap=True)
        _cell(ws, r,  9, o.get("technicien",""),              bg=bg)
        _cell(ws, r, 10, arret,
              bg=LIGHT_RED if arret > 8 else bg,
              fg=RED if arret > 8 else DARK_GRAY,
              bold=arret > 8, h="right", fmt="0.0")
        _cell(ws, r, 11, float(o.get("duree_intervention") or 0),
              bg=bg, h="right", fmt="0.0")
        _cell(ws, r, 12, o.get("pieces_utilisees",""),        bg=bg, wrap=True)
        _cell(ws, r, 13, float(o.get("cout_mo") or 0),        bg=bg, h="right", fmt="#,##0")
        _cell(ws, r, 14, float(o.get("cout_pieces") or 0),    bg=bg, h="right", fmt="#,##0")
        _cell(ws, r, 15, cout, bg=bg, fg=DARK_GRAY, bold=True, h="right", fmt="#,##0")
        _cell(ws, r, 16, st,   bg=sbg, fg=sfg, bold=True, h="center")
        _cell(ws, r, 17, o.get("observations",""), bg=bg, wrap=True)

    # Ligne totaux
    last = 4 + len(interventions) + 1
    for col in range(1, 18):
        ws.cell(row=last, column=col).fill = _fill(GOLD)
        ws.cell(row=last, column=col).font = Font(bold=True, size=10, color=NAVY, name="Calibri")
        ws.cell(row=last, column=col).alignment = _align(h="right")
        ws.cell(row=last, column=col).border = _border()
    ws.cell(row=last, column=1).value = "TOTAUX"
    ws.cell(row=last, column=1).alignment = _align(h="left")
    ws.cell(row=last, column=10).value = t_arret;  ws.cell(row=last, column=10).number_format = "0.0"
    ws.cell(row=last, column=13).value = t_mo;     ws.cell(row=last, column=13).number_format = "#,##0"
    ws.cell(row=last, column=14).value = t_pi;     ws.cell(row=last, column=14).number_format = "#,##0"
    ws.cell(row=last, column=15).value = t_tot;    ws.cell(row=last, column=15).number_format = "#,##0"
    ws.row_dimensions[last].height = 20

    _col_w(ws, {"A":14,"B":14,"C":13,"D":12,"E":24,"F":22,
                "G":28,"H":38,"I":20,"J":10,"K":12,"L":28,
                "M":18,"N":18,"O":18,"P":12,"Q":20})
    ws.freeze_panes = "E5"
    ws.auto_filter.ref = f"A4:Q{last-1}"


# ── FEUILLE 5 : PIÈCES DE RECHANGE ───────────────────────────────────────────
def _sheet_pieces(wb, pieces: list) -> None:
    ws = wb.create_sheet("PIÈCES DE RECHANGE")
    ws.sheet_view.showGridLines = False
    ws.sheet_properties.tabColor = PURPLE

    val_stock = sum(float(p.get("stock_actuel") or 0) * float(p.get("prix_unitaire") or 0)
                    for p in pieces)
    nb_alerte = sum(1 for p in pieces
                    if float(p.get("stock_actuel") or 0) <= float(p.get("stock_min") or 0))

    _title_row(ws, 1, f"GESTION DES PIÈCES DE RECHANGE — {YEAR}", 14)
    _sub_row(ws, 2,
             f"{len(pieces)} références   |   "
             f"Valeur stock : {val_stock:,.0f} FCFA   |   "
             f"Alertes : {nb_alerte}   |   "
             f"Criticité A : {sum(1 for p in pieces if p.get('criticite')=='A')}", 14)
    ws.row_dimensions[3].height = 6

    hdr = ["Criticité","Code","Désignation","Réf. Fabricant","Équipements",
           "Unité","Stock actuel","Stock min","Stock max","Emplacement",
           "Prix unitaire (FCFA)","Valeur stock (FCFA)","Fournisseur",
           "Délai appro (j)","Observations"]
    _header(ws, 4, hdr, bg=NAVY, fg=PURPLE)

    for i, p in enumerate(pieces):
        r    = 5 + i
        stk  = float(p.get("stock_actuel") or 0)
        mn   = float(p.get("stock_min") or 0)
        mx   = float(p.get("stock_max") or 1)
        pu   = float(p.get("prix_unitaire") or 0)
        val  = stk * pu
        crit = str(p.get("criticite","B"))
        cbg, cfg = CRIT_BG.get(crit, (LIGHT_GRAY, DARK_GRAY)), CRIT_FG.get(crit, DARK_GRAY)
        alerte = stk <= mn
        bg = LIGHT_RED if stk == 0 else (LIGHT_ORANGE if alerte else (LIGHT_BLUE if i%2==0 else WHITE))

        _cell(ws, r,  1, crit, bg=cbg, fg=cfg, bold=True, h="center")
        _cell(ws, r,  2, p.get("code_piece",""),           bg=bg, fg=DARK_GRAY, bold=True)
        _cell(ws, r,  3, p.get("designation",""),          bg=bg)
        _cell(ws, r,  4, p.get("reference_fabricant",""),  bg=bg)
        _cell(ws, r,  5, p.get("equipements_concernes",""),bg=bg, wrap=True)
        _cell(ws, r,  6, p.get("unite",""),                bg=bg, h="center")
        # Stock actuel coloré
        stk_bg = LIGHT_RED if alerte else bg
        stk_fg = RED if alerte else DARK_GRAY
        _cell(ws, r,  7, stk,  bg=stk_bg, fg=stk_fg, bold=alerte, h="right", fmt="#,##0.0")
        _cell(ws, r,  8, mn,   bg=bg, h="right", fmt="#,##0.0")
        _cell(ws, r,  9, mx,   bg=bg, h="right", fmt="#,##0.0")
        _cell(ws, r, 10, p.get("emplacement_magasin",""), bg=bg)
        _cell(ws, r, 11, pu,   bg=bg, h="right", fmt="#,##0")
        _cell(ws, r, 12, val,  bg=bg, fg=DARK_GRAY, bold=True, h="right", fmt="#,##0")
        _cell(ws, r, 13, p.get("fournisseur",""),          bg=bg)
        _cell(ws, r, 14, int(p.get("delai_appro") or 0),   bg=bg, h="center")
        _cell(ws, r, 15, p.get("observations",""),         bg=bg, wrap=True)

    # Totaux
    last = 4 + len(pieces) + 1
    ws.merge_cells(f"A{last}:K{last}")
    c = ws[f"A{last}"]
    c.value = "VALEUR TOTALE STOCK"
    c.fill = _fill(GOLD); c.font = Font(bold=True, size=10, color=NAVY, name="Calibri")
    c.alignment = _align(h="right")
    _cell(ws, last, 12, val_stock, bg=GOLD, fg=NAVY, bold=True, h="right", fmt="#,##0")

    _col_w(ws, {"A":9,"B":12,"C":32,"D":20,"E":22,"F":8,
                "G":12,"H":10,"I":10,"J":16,"K":20,"L":22,
                "M":22,"N":12,"O":25})
    ws.freeze_panes = "B5"
    ws.auto_filter.ref = f"A4:O{last-1}"


# ── FEUILLE 6 : PRESTATAIRES ──────────────────────────────────────────────────
def _sheet_prestataires(wb, prestataires: list) -> None:
    ws = wb.create_sheet("PRESTATAIRES")
    ws.sheet_view.showGridLines = False
    ws.sheet_properties.tabColor = TEAL

    val_contrats = sum(float(p.get("montant_contrat") or 0) for p in prestataires
                       if p.get("contrat") == "Oui")

    _title_row(ws, 1, f"FOURNISSEURS & PRESTATAIRES DE MAINTENANCE — {YEAR}", 13)
    _sub_row(ws, 2,
             f"{len(prestataires)} partenaires   |   "
             f"Contrats actifs : {sum(1 for p in prestataires if p.get('contrat')=='Oui')}   |   "
             f"Valeur contrats : {val_contrats:,.0f} FCFA/an", 13)
    ws.row_dimensions[3].height = 6

    hdr = ["Code","Raison Sociale","Spécialité","Contact","Téléphone","Email",
           "Contrat","Type contrat","Début","Fin",
           "Montant/an (FCFA)","Note qualité","Observations"]
    _header(ws, 4, hdr, bg=NAVY, fg=TEAL)

    for i, p in enumerate(prestataires):
        r   = 5 + i
        bg  = LIGHT_BLUE if i % 2 == 0 else WHITE
        cont = str(p.get("contrat","Non"))
        cbg, cfg = (LIGHT_GREEN, GREEN) if cont == "Oui" else (LIGHT_GRAY, DARK_GRAY)
        note = float(p.get("note_qualite") or 0)
        stars = "★" * int(note) + "☆" * (5 - int(note))
        note_fg = GOLD if note >= 4.5 else (ORANGE if note >= 3.5 else RED)

        _cell(ws, r,  1, p.get("code_prestataire",""), bg=bg, fg=BLUE, bold=True)
        _cell(ws, r,  2, p.get("raison_sociale",""),   bg=bg)
        _cell(ws, r,  3, p.get("specialite",""),       bg=bg)
        _cell(ws, r,  4, p.get("contact",""),          bg=bg)
        _cell(ws, r,  5, p.get("telephone",""),        bg=bg)
        _cell(ws, r,  6, p.get("email",""),            bg=bg)
        _cell(ws, r,  7, cont, bg=cbg, fg=cfg, bold=True, h="center")
        _cell(ws, r,  8, p.get("type_contrat",""),     bg=bg, h="center")
        _cell(ws, r,  9, p.get("debut_contrat",""),    bg=bg, h="center")
        _cell(ws, r, 10, p.get("fin_contrat",""),      bg=bg, h="center")
        _cell(ws, r, 11, float(p.get("montant_contrat") or 0),
              bg=bg, h="right", bold=True, fg=DARK_GRAY, fmt="#,##0")
        # Étoiles note qualité
        c_note = ws.cell(row=r, column=12, value=f"{note:.1f}  {stars}")
        c_note.fill = _fill(bg)
        c_note.font = Font(size=12, color=note_fg, bold=True, name="Calibri")
        c_note.alignment = _align(h="center"); c_note.border = _border()
        ws.row_dimensions[r].height = 17
        _cell(ws, r, 13, p.get("observations",""), bg=bg, wrap=True)

    _col_w(ws, {"A":14,"B":28,"C":26,"D":18,"E":18,
                "F":28,"G":10,"H":16,"I":12,"J":12,
                "K":22,"L":18,"M":25})
    ws.freeze_panes = "B5"
    ws.auto_filter.ref = f"A4:M{4+len(prestataires)}"


# ── FEUILLE 7 : INDICATEURS ───────────────────────────────────────────────────
def _sheet_indicateurs(wb, indicateurs: list, dash: dict) -> None:
    ws = wb.create_sheet("INDICATEURS")
    ws.sheet_view.showGridLines = False
    ws.sheet_properties.tabColor = RED

    _title_row(ws, 1, f"INDICATEURS DE PERFORMANCE MAINTENANCE (KPIs) — {YEAR}", 14)
    _sub_row(ws, 2,
             "MTBF = Temps moyen entre pannes   |   MTTR = Temps moyen de réparation   |   "
             "Disponibilité = MTBF / (MTBF + MTTR)   |   Objectif dispo ≥ 95%", 14)
    ws.row_dimensions[3].height = 6

    # Définitions
    _title_row(ws, 4, "DÉFINITIONS DES INDICATEURS CLÉS", 14, bg=STEEL, size=11)
    ws.row_dimensions[4].height = 22
    defs = [
        ("MTBF (h)",         "Mean Time Between Failures = Temps de marche total / Nb pannes. Objectif criticité A : > 300 h"),
        ("MTTR (h)",         "Mean Time To Repair = Σ Temps arrêt / Nb pannes. Objectif criticité A : < 4 h"),
        ("Disponibilité %",  "= MTBF / (MTBF + MTTR) × 100. Objectif global ≥ 95%. Criticité A ≥ 97%"),
        ("Ratio PM %",       "= Nb OT Préventifs / Nb OT total × 100. Objectif : ≥ 70% (maintenance proactive)"),
        ("Coût maintenance", "Coût MO + Pièces. Benchmark : 2 à 4% de la valeur du parc par an"),
        ("OEE %",            "Overall Equipment Effectiveness = Disponibilité × Performance × Qualité. Classe monde : OEE > 85%"),
    ]
    for i, (kpi, defn) in enumerate(defs):
        r = 5 + i
        c1 = ws.cell(row=r, column=1, value=kpi)
        c1.fill = _fill(STEEL); c1.font = Font(bold=True, size=9, color=GOLD, name="Calibri")
        c1.alignment = _align(h="center", v="center"); c1.border = _border()
        ws.merge_cells(f"B{r}:N{r}")
        c2 = ws.cell(row=r, column=2, value=defn)
        c2.fill = _fill(LIGHT_BLUE if i%2==0 else WHITE)
        c2.font = Font(size=9, color=DARK_GRAY, italic=True, name="Calibri")
        c2.alignment = _align(h="left", v="center"); c2.border = _border()
        ws.row_dimensions[r].height = 20

    ws.row_dimensions[11].height = 8

    # Tableau MTBF/MTTR par équipement
    _title_row(ws, 12, f"MTBF / MTTR / DISPONIBILITÉ PAR ÉQUIPEMENT — {YEAR}", 14,
               bg=NAVY, size=11)
    ws.row_dimensions[12].height = 22

    hdr_ind = ["Criticité","Code équipement","Désignation","Famille",
               "Nb pannes YTD","Arrêt total (h)","MTBF (h)","MTTR (h)",
               "Disponibilité %","Objectif %","Écart"]
    _header(ws, 13, hdr_ind, bg=NAVY, fg=GOLD)

    OBJ = 95.0
    for i, r_d in enumerate(indicateurs):
        r    = 14 + i
        crit = str(r_d.get("criticite","B"))
        cbg, cfg = CRIT_BG.get(crit, (LIGHT_GRAY, DARK_GRAY)), CRIT_FG.get(crit, DARK_GRAY)
        dispo   = float(r_d.get("disponibilite") or 100)
        ecart   = dispo - OBJ
        dispo_bg = LIGHT_GREEN if dispo >= OBJ else (LIGHT_ORANGE if dispo >= 90 else LIGHT_RED)
        dispo_fg = GREEN if dispo >= OBJ else (ORANGE if dispo >= 90 else RED)
        nb      = int(r_d.get("nb_pannes") or 0)
        bg      = LIGHT_BLUE if i % 2 == 0 else WHITE

        _cell(ws, r,  1, crit, bg=cbg, fg=cfg, bold=True, h="center")
        _cell(ws, r,  2, r_d.get("code_equipement",""), bg=bg, fg=BLUE, bold=True)
        _cell(ws, r,  3, r_d.get("designation",""),     bg=bg)
        _cell(ws, r,  4, r_d.get("famille",""),         bg=bg)
        _cell(ws, r,  5, nb,  bg=LIGHT_RED if nb > 3 else bg,
              fg=RED if nb > 3 else DARK_GRAY, bold=nb > 3, h="center")
        _cell(ws, r,  6, float(r_d.get("total_arret") or 0),
              bg=bg, fg=RED, h="right", fmt="0.0")
        _cell(ws, r,  7, float(r_d.get("mtbf") or 0),
              bg=LIGHT_GREEN if float(r_d.get("mtbf") or 0) >= 300 else LIGHT_ORANGE,
              fg=GREEN if float(r_d.get("mtbf") or 0) >= 300 else ORANGE,
              bold=True, h="right", fmt="0")
        _cell(ws, r,  8, float(r_d.get("mttr") or 0),
              bg=LIGHT_RED if float(r_d.get("mttr") or 0) > 8 else bg,
              fg=RED if float(r_d.get("mttr") or 0) > 8 else ORANGE,
              bold=True, h="right", fmt="0.0")
        _cell(ws, r,  9, dispo / 100,
              bg=dispo_bg, fg=dispo_fg, bold=True, h="center", fmt="0.0%")
        _cell(ws, r, 10, OBJ / 100,
              bg=bg, fg=DARK_GRAY, h="center", fmt="0%")
        _cell(ws, r, 11, ecart / 100,
              bg=LIGHT_GREEN if ecart >= 0 else LIGHT_RED,
              fg=GREEN if ecart >= 0 else RED,
              bold=True, h="center", fmt="+0.0%;-0.0%;0.0%")

    if not indicateurs:
        ws.merge_cells("A14:K14")
        c = ws["A14"]
        c.value = "Aucune intervention enregistrée — saisir des OT pour voir les indicateurs."
        c.fill = _fill(LIGHT_BLUE)
        c.font = Font(size=10, color=BLUE, italic=True, name="Calibri")
        c.alignment = _align(h="center")

    # Synthèse mensuelle (depuis dashboard)
    monthly = dash.get("monthly", [])
    if monthly:
        row_m = 14 + max(len(indicateurs), 1) + 2
        ws.row_dimensions[row_m - 1].height = 8
        _title_row(ws, row_m, f"SUIVI MENSUEL DES INTERVENTIONS — {YEAR}", 14,
                   bg=STEEL, size=11)
        ws.row_dimensions[row_m].height = 22
        hdr_m = ["Mois","Nb OT PM","Nb OT CM","Arrêts CM (h)",
                 "Coût PM (FCFA)","Coût CM (FCFA)","Coût total (FCFA)"]
        _header(ws, row_m + 1, hdr_m, bg=STEEL, fg=GOLD)
        t_pm = t_cm = t_arr = t_cpm = t_ccm = 0
        for j, m in enumerate(monthly):
            r = row_m + 2 + j
            nb_pm = int(m.get("nb_pm") or 0)
            nb_cm = int(m.get("nb_cm") or 0)
            h_arr = float(m.get("h_arret") or 0)
            cout  = float(m.get("cout") or 0)
            has   = nb_pm > 0 or nb_cm > 0
            bg    = LIGHT_BLUE if j % 2 == 0 else WHITE
            _cell(ws, r, 1, m["mois"],                   bg=bg, bold=True)
            _cell(ws, r, 2, nb_pm if has else "",         bg=LIGHT_GREEN if has else bg, fg=GREEN if has else DARK_GRAY, h="center")
            _cell(ws, r, 3, nb_cm if has else "",         bg=LIGHT_RED if (has and nb_cm>0) else bg, fg=RED if nb_cm>0 else DARK_GRAY, h="center")
            _cell(ws, r, 4, h_arr if has else "",         bg=LIGHT_ORANGE if (has and h_arr>0) else bg, h="right", fmt="0.0")
            _cell(ws, r, 5, "" , bg=bg, h="right", fmt="#,##0")
            _cell(ws, r, 6, cout if has else "", bg=bg, h="right", fmt="#,##0")
            _cell(ws, r, 7, cout if has else "", bg=bg if not has else LIGHT_BLUE, bold=has, h="right", fmt="#,##0")
            t_pm += nb_pm; t_cm += nb_cm; t_arr += h_arr; t_ccm += cout

        last_m = row_m + 2 + len(monthly)
        for col in range(1, 8):
            ws.cell(row=last_m, column=col).fill = _fill(GOLD)
            ws.cell(row=last_m, column=col).font = Font(bold=True, size=10, color=NAVY, name="Calibri")
            ws.cell(row=last_m, column=col).alignment = _align(h="right")
            ws.cell(row=last_m, column=col).border = _border()
        ws.cell(row=last_m, column=1).value = "TOTAL YTD"
        ws.cell(row=last_m, column=2).value = t_pm
        ws.cell(row=last_m, column=3).value = t_cm
        ws.cell(row=last_m, column=4).value = t_arr;  ws.cell(row=last_m, column=4).number_format = "0.0"
        ws.cell(row=last_m, column=6).value = t_ccm;  ws.cell(row=last_m, column=6).number_format = "#,##0"
        ws.cell(row=last_m, column=7).value = t_ccm;  ws.cell(row=last_m, column=7).number_format = "#,##0"
        ws.row_dimensions[last_m].height = 20

    _col_w(ws, {"A":12,"B":20,"C":28,"D":20,
                "E":14,"F":16,"G":14,"H":12,"I":16,"J":12,"K":14,
                "L":14,"M":14,"N":14})
    ws.freeze_panes = "B14"
